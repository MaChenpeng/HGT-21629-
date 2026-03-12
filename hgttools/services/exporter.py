"""Excel导出服务"""

from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from ..models.material import Part, SummaryItem


class ExcelExporter:
    """Excel导出器"""

    # 标题行样式
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 边框样式
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 小计行样式
    SUBTOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    SUBTOTAL_FONT = Font(bold=True)

    def __init__(self):
        self.wb = None

    def export(self, parts: List[Part], summary_items: List[SummaryItem],
               single_support_parts: List[Part] = None, output_path: str = None) -> str:
        """
        导出BOM清单到Excel

        Args:
            parts: 零件明细列表
            summary_items: 汇总列表
            single_support_parts: 单个支架表零件列表
            output_path: 输出文件路径

        Returns:
            输出文件路径
        """
        self.wb = Workbook()

        # 创建汇总表工作表
        self._create_summary_sheet(summary_items)

        # 创建零件明细表
        if parts:
            self._create_parts_sheet(parts)

        # 创建单个支架表
        if single_support_parts:
            self._create_single_support_sheet(single_support_parts)

        # 保存文件
        if output_path:
            self.wb.save(output_path)
            return output_path

        return None

    def _create_summary_sheet(self, items: List[SummaryItem]):
        """创建汇总表工作表"""
        # 使用第一个工作表或创建新表
        ws = self.wb.active
        ws.title = "汇总表"

        # 设置标题行
        headers = [
            "序号", "名称及编码", "规格", "等级", "设计量", "设计裕量",
            "总量", "单位", "单重(Kg)", "总重(Kg)", "技术条件", "备注",
            "修改", "材质", "单位表面积m/m²", "零件面积m²"
        ]

        # 写入标题行
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # 写入数据行
        for row_idx, item in enumerate(items, 2):
            self._write_summary_row(ws, row_idx, item)

        # 设置列宽
        column_widths = {
            'A': 6,   # 序号
            'B': 20,  # 名称及编码
            'C': 15,  # 规格
            'D': 8,   # 等级
            'E': 10,  # 设计量
            'F': 10,  # 设计裕量
            'G': 10,  # 总量
            'H': 6,   # 单位
            'I': 12,  # 单重
            'J': 12,  # 总重
            'K': 25,  # 技术条件
            'L': 10,  # 备注
            'M': 6,   # 修改
            'N': 12,  # 材质
            'O': 15,  # 单位表面积
            'P': 12,  # 零件面积
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 冻结首行
        ws.freeze_panes = 'A2'

    def _write_summary_row(self, ws, row_idx: int, item: SummaryItem):
        """写入汇总表数据行"""
        # 设计量（E列）和总量（G列）相同
        design_qty = item.design_quantity if item.design_quantity else 0

        data = [
            item.seq_no,                    # A: 序号
            item.name,                      # B: 名称及编码
            item.specification,             # C: 规格
            item.grade or "",               # D: 等级
            design_qty,                     # E: 设计量
            item.margin or 0,               # F: 设计裕量
            f"=E{row_idx}+F{row_idx}",      # G: 总量 (公式)
            item.unit,                      # H: 单位
            item.unit_weight,               # I: 单重
            f"=G{row_idx}*I{row_idx}",      # J: 总重 (公式)
            item.tech_condition,            # K: 技术条件
            item.remark,                    # L: 备注
            item.grade or "",               # M: 修改
            item.material,                  # N: 材质
            item.unit_surface_area,         # O: 单位表面积
            f"=G{row_idx}*O{row_idx}",      # P: 零件面积 (公式)
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # 设置值
            if isinstance(value, str) and value.startswith('='):
                cell.value = value  # 公式
            else:
                cell.value = value

            # 设置边框
            cell.border = self.THIN_BORDER

            # 数字右对齐
            if col_idx in [1, 5, 6, 7, 9, 10, 15, 16]:
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')

    def _create_parts_sheet(self, parts: List[Part]):
        """创建零件明细表"""
        ws = self.wb.create_sheet(title="零件明细表")

        # 设置标题行
        headers = [
            "序号", "管架编号", "名称", "规格", "等级", "数量",
            "单位", "材质", "单重(Kg)", "总重(Kg)", "技术条件",
            "备注", "管架序号", "单位表面积m/m²", "零件面积m²"
        ]

        # 写入标题行
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # 写入数据
        for row_idx, part in enumerate(parts, 2):
            self._write_part_row(ws, row_idx, part)

        # 设置列宽
        column_widths = {
            'A': 6,   # 序号
            'B': 20,  # 管架编号
            'C': 15,  # 名称
            'D': 15,  # 规格
            'E': 8,   # 等级
            'F': 10,  # 数量
            'G': 6,   # 单位
            'H': 12,  # 材质
            'I': 12,  # 单重
            'J': 12,  # 总重
            'K': 25,  # 技术条件
            'L': 10,  # 备注
            'M': 8,   # 管架序号
            'N': 15,  # 单位表面积
            'O': 12,  # 零件面积
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 冻结首行
        ws.freeze_panes = 'A2'

    def _write_part_row(self, ws, row_idx: int, part: Part):
        """写入零件明细行"""
        data = [
            part.seq_no,                # A: 序号
            part.support_code,          # B: 管架编号
            part.name,                  # C: 名称
            part.specification,         # D: 规格
            part.grade or "",           # E: 等级
            part.quantity,              # F: 数量
            part.unit,                  # G: 单位
            part.material,              # H: 材质
            part.unit_weight,           # I: 单重
            part.total_weight,          # J: 总重
            part.tech_condition,        # K: 技术条件
            part.remark,                # L: 备注
            part.support_seq,           # M: 管架序号
            part.unit_surface_area,     # N: 单位表面积
            part.surface_area,          # O: 零件面积
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = self.THIN_BORDER

            # 数字右对齐
            if col_idx in [1, 6, 9, 10, 14, 15]:
                cell.alignment = Alignment(horizontal='right')

    def _create_single_support_sheet(self, parts: List[Part]):
        """创建单个支架表"""
        ws = self.wb.create_sheet(title="单个支架表")

        # 设置标题行（与零件表类似）
        headers = [
            "序号", "管架编号", "名称", "规格", "等级", "数量",
            "单位", "材质", "单重(Kg)", "总重(Kg)", "技术条件",
            "备注", "单位表面积m/m²", "零件面积m²"
        ]

        # 写入标题行
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # 写入数据
        for row_idx, part in enumerate(parts, 2):
            data = [
                part.seq_no,
                part.support_code,
                part.name,
                part.specification,
                part.grade or "",
                part.quantity,
                part.unit,
                part.material,
                part.unit_weight,
                part.total_weight,
                part.tech_condition,
                part.remark,
                part.unit_surface_area,
                part.surface_area,
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.THIN_BORDER

        # 设置列宽
        column_widths = [6, 20, 15, 15, 8, 10, 6, 12, 12, 12, 25, 10, 15, 12]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # 冻结首行
        ws.freeze_panes = 'A2'

    def add_subtotals(self, ws_name: str = "汇总表"):
        """
        在汇总表中添加小计行（按名称分组）
        这是一个可选功能
        """
        if ws_name not in self.wb.sheetnames:
            return

        ws = self.wb[ws_name]

        # 遍历数据行，按名称分组插入小计
        current_name = None
        group_start_row = 2
        rows_to_insert = []

        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=2).value  # B列：名称

            if name != current_name and current_name is not None:
                # 记录需要插入小计的位置
                rows_to_insert.append((row, current_name, group_start_row, row - 1))
                group_start_row = row

            current_name = name

        # 从后往前插入小计行（避免行号变化）
        for insert_row, name, start_row, end_row in reversed(rows_to_insert):
            ws.insert_rows(insert_row)

            # 设置小计行内容
            ws.cell(row=insert_row, column=2, value="小 计").font = self.SUBTOTAL_FONT
            ws.cell(row=insert_row, column=8, value=name)  # 材质

            # 设置公式
            ws.cell(row=insert_row, column=5, value=f"=SUM(E{start_row}:E{end_row})")
            ws.cell(row=insert_row, column=10, value=f"=SUM(J{start_row}:J{end_row})")

            # 设置样式
            for col in range(1, 17):
                cell = ws.cell(row=insert_row, column=col)
                cell.fill = self.SUBTOTAL_FILL
                cell.border = self.THIN_BORDER
